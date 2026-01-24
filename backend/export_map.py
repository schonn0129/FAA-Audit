"""
Export functionality for FAA Audit data.
Supports CSV and Excel formats.
"""

import csv
import io
from typing import List, Dict, Any


def export_questions_to_csv(questions: List[Dict[str, Any]]) -> str:
    """
    Export questions to CSV format.

    Args:
        questions: List of question dictionaries

    Returns:
        CSV string
    """
    output = io.StringIO()

    # Define column headers matching the expected CSV format
    fieldnames = [
        'Element_ID',
        'QID',
        'Question_Number',
        'Question_Text_Full',
        'Question_Text_Condensed',
        'Data_Collection_Guidance',
        'Reference_Raw',
        'Reference_CFR_List',
        'Reference_FAA_Guidance_List',
        'Reference_Other_List',
        'PDF_Page_Number',
        'PDF_Element_Block_ID',
        'Notes'
    ]

    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction='ignore')
    writer.writeheader()

    for q in questions:
        row = {
            'Element_ID': q.get('Element_ID', ''),
            'QID': q.get('QID', ''),
            'Question_Number': q.get('Question_Number', ''),
            'Question_Text_Full': q.get('Question_Text_Full', ''),
            'Question_Text_Condensed': q.get('Question_Text_Condensed', ''),
            'Data_Collection_Guidance': q.get('Data_Collection_Guidance', ''),
            'Reference_Raw': q.get('Reference_Raw', ''),
            'Reference_CFR_List': '; '.join(q.get('Reference_CFR_List', [])) if isinstance(q.get('Reference_CFR_List'), list) else q.get('Reference_CFR_List', ''),
            'Reference_FAA_Guidance_List': '; '.join(q.get('Reference_FAA_Guidance_List', [])) if isinstance(q.get('Reference_FAA_Guidance_List'), list) else q.get('Reference_FAA_Guidance_List', ''),
            'Reference_Other_List': '; '.join(q.get('Reference_Other_List', [])) if isinstance(q.get('Reference_Other_List'), list) else q.get('Reference_Other_List', ''),
            'PDF_Page_Number': q.get('PDF_Page_Number', ''),
            'PDF_Element_Block_ID': q.get('PDF_Element_Block_ID', ''),
            'Notes': '; '.join(q.get('Notes', [])) if isinstance(q.get('Notes'), list) else q.get('Notes', '')
        }
        writer.writerow(row)

    return output.getvalue()


def export_map_to_csv(map_rows: List[Dict[str, Any]]) -> str:
    """
    Export MAP rows to CSV format.

    Args:
        map_rows: List of MAP row dictionaries

    Returns:
        CSV string
    """
    output = io.StringIO()

    fieldnames = [
        'QID',
        'Question_Text',
        'AIP_Reference',
        'GMM_Reference',
        'Other_Manual_References',
        'Evidence_Required',
        'Applicability_Status',
        'Applicability_Reason',
        'Audit_Finding',
        'Compliance_Status'
    ]

    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction='ignore')
    writer.writeheader()

    for row in map_rows:
        writer.writerow({
            'QID': row.get('QID', ''),
            'Question_Text': row.get('Question_Text', ''),
            'AIP_Reference': row.get('AIP_Reference', ''),
            'GMM_Reference': row.get('GMM_Reference', ''),
            'Other_Manual_References': row.get('Other_Manual_References', ''),
            'Evidence_Required': row.get('Evidence_Required', ''),
            'Applicability_Status': row.get('Applicability_Status', ''),
            'Applicability_Reason': row.get('Applicability_Reason', ''),
            'Audit_Finding': row.get('Audit_Finding', ''),
            'Compliance_Status': row.get('Compliance_Status', '')
        })

    return output.getvalue()


def export_audit_to_csv(audit: Dict[str, Any]) -> str:
    """
    Export a single audit's questions to CSV format.

    Args:
        audit: Audit dictionary containing questions

    Returns:
        CSV string
    """
    questions = audit.get('questions', [])
    return export_questions_to_csv(questions)


def export_audits_to_csv(audits: List[Dict[str, Any]]) -> str:
    """
    Export multiple audits to CSV format.
    Includes audit metadata as additional columns.

    Args:
        audits: List of audit dictionaries

    Returns:
        CSV string
    """
    output = io.StringIO()

    fieldnames = [
        'Audit_ID',
        'Audit_Filename',
        'Element_ID',
        'QID',
        'Question_Number',
        'Question_Text_Full',
        'Question_Text_Condensed',
        'Data_Collection_Guidance',
        'Reference_Raw',
        'Reference_CFR_List',
        'Reference_FAA_Guidance_List',
        'Reference_Other_List',
        'PDF_Page_Number',
        'PDF_Element_Block_ID',
        'Notes'
    ]

    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction='ignore')
    writer.writeheader()

    for audit in audits:
        audit_id = audit.get('id', '')
        filename = audit.get('filename', '')
        questions = audit.get('questions', [])

        for q in questions:
            row = {
                'Audit_ID': audit_id,
                'Audit_Filename': filename,
                'Element_ID': q.get('Element_ID', ''),
                'QID': q.get('QID', ''),
                'Question_Number': q.get('Question_Number', ''),
                'Question_Text_Full': q.get('Question_Text_Full', ''),
                'Question_Text_Condensed': q.get('Question_Text_Condensed', ''),
                'Data_Collection_Guidance': q.get('Data_Collection_Guidance', ''),
                'Reference_Raw': q.get('Reference_Raw', ''),
                'Reference_CFR_List': '; '.join(q.get('Reference_CFR_List', [])) if isinstance(q.get('Reference_CFR_List'), list) else q.get('Reference_CFR_List', ''),
                'Reference_FAA_Guidance_List': '; '.join(q.get('Reference_FAA_Guidance_List', [])) if isinstance(q.get('Reference_FAA_Guidance_List'), list) else q.get('Reference_FAA_Guidance_List', ''),
                'Reference_Other_List': '; '.join(q.get('Reference_Other_List', [])) if isinstance(q.get('Reference_Other_List'), list) else q.get('Reference_Other_List', ''),
                'PDF_Page_Number': q.get('PDF_Page_Number', ''),
                'PDF_Element_Block_ID': q.get('PDF_Element_Block_ID', ''),
                'Notes': '; '.join(q.get('Notes', [])) if isinstance(q.get('Notes'), list) else q.get('Notes', '')
            }
            writer.writerow(row)

    return output.getvalue()


def export_questions_to_xlsx(questions: List[Dict[str, Any]]) -> bytes:
    """
    Export questions to Excel format.

    Args:
        questions: List of question dictionaries

    Returns:
        Excel file bytes
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Questions"

    # Define headers
    headers = [
        'Element_ID',
        'QID',
        'Question_Number',
        'Question_Text_Full',
        'Question_Text_Condensed',
        'Data_Collection_Guidance',
        'Reference_Raw',
        'Reference_CFR_List',
        'Reference_FAA_Guidance_List',
        'Reference_Other_List',
        'PDF_Page_Number',
        'PDF_Element_Block_ID',
        'Notes'
    ]

    # Style for headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Write data
    for row_idx, q in enumerate(questions, 2):
        ws.cell(row=row_idx, column=1, value=q.get('Element_ID', ''))
        ws.cell(row=row_idx, column=2, value=q.get('QID', ''))
        ws.cell(row=row_idx, column=3, value=q.get('Question_Number', ''))
        ws.cell(row=row_idx, column=4, value=q.get('Question_Text_Full', ''))
        ws.cell(row=row_idx, column=5, value=q.get('Question_Text_Condensed', ''))
        ws.cell(row=row_idx, column=6, value=q.get('Data_Collection_Guidance', ''))
        ws.cell(row=row_idx, column=7, value=q.get('Reference_Raw', ''))

        cfr_list = q.get('Reference_CFR_List', [])
        ws.cell(row=row_idx, column=8, value='; '.join(cfr_list) if isinstance(cfr_list, list) else cfr_list)

        faa_list = q.get('Reference_FAA_Guidance_List', [])
        ws.cell(row=row_idx, column=9, value='; '.join(faa_list) if isinstance(faa_list, list) else faa_list)

        other_list = q.get('Reference_Other_List', [])
        ws.cell(row=row_idx, column=10, value='; '.join(other_list) if isinstance(other_list, list) else other_list)

        ws.cell(row=row_idx, column=11, value=q.get('PDF_Page_Number', ''))
        ws.cell(row=row_idx, column=12, value=q.get('PDF_Element_Block_ID', ''))

        notes = q.get('Notes', [])
        ws.cell(row=row_idx, column=13, value='; '.join(notes) if isinstance(notes, list) else notes)

    # Adjust column widths
    column_widths = [12, 12, 15, 80, 40, 50, 30, 30, 30, 20, 15, 20, 40]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_map_to_xlsx(map_rows: List[Dict[str, Any]]) -> bytes:
    """
    Export MAP rows to Excel format.

    Args:
        map_rows: List of MAP row dictionaries

    Returns:
        Excel file bytes
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MAP"

    headers = [
        'QID',
        'Question_Text',
        'AIP_Reference',
        'GMM_Reference',
        'Other_Manual_References',
        'Evidence_Required',
        'Applicability_Status',
        'Applicability_Reason',
        'Audit_Finding',
        'Compliance_Status'
    ]

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    header_align = Alignment(wrap_text=True, vertical="top")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, row in enumerate(map_rows, 2):
        ws.cell(row=row_idx, column=1, value=row.get('QID', ''))
        ws.cell(row=row_idx, column=2, value=row.get('Question_Text', ''))
        ws.cell(row=row_idx, column=3, value=row.get('AIP_Reference', ''))
        ws.cell(row=row_idx, column=4, value=row.get('GMM_Reference', ''))
        ws.cell(row=row_idx, column=5, value=row.get('Other_Manual_References', ''))
        ws.cell(row=row_idx, column=6, value=row.get('Evidence_Required', ''))
        ws.cell(row=row_idx, column=7, value=row.get('Applicability_Status', ''))
        ws.cell(row=row_idx, column=8, value=row.get('Applicability_Reason', ''))
        ws.cell(row=row_idx, column=9, value=row.get('Audit_Finding', ''))
        ws.cell(row=row_idx, column=10, value=row.get('Compliance_Status', ''))

    column_widths = [12, 80, 25, 25, 30, 45, 20, 35, 30, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_audit_to_xlsx(audit: Dict[str, Any]) -> bytes:
    """
    Export a single audit to Excel format.

    Args:
        audit: Audit dictionary

    Returns:
        Excel file bytes
    """
    questions = audit.get('questions', [])
    return export_questions_to_xlsx(questions)


def export_audits_to_xlsx(audits: List[Dict[str, Any]]) -> bytes:
    """
    Export multiple audits to Excel format with multiple sheets.

    Args:
        audits: List of audit dictionaries

    Returns:
        Excel file bytes
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    headers = [
        'Element_ID',
        'QID',
        'Question_Number',
        'Question_Text_Full',
        'Question_Text_Condensed',
        'Data_Collection_Guidance',
        'Reference_Raw',
        'Reference_CFR_List',
        'Reference_FAA_Guidance_List',
        'Reference_Other_List',
        'PDF_Page_Number',
        'PDF_Element_Block_ID',
        'Notes'
    ]

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    for audit in audits:
        # Create sheet with filename (truncated if needed)
        filename = audit.get('filename', 'Unknown')[:30]
        ws = wb.create_sheet(title=filename)

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Write questions
        questions = audit.get('questions', [])
        for row_idx, q in enumerate(questions, 2):
            ws.cell(row=row_idx, column=1, value=q.get('Element_ID', ''))
            ws.cell(row=row_idx, column=2, value=q.get('QID', ''))
            ws.cell(row=row_idx, column=3, value=q.get('Question_Number', ''))
            ws.cell(row=row_idx, column=4, value=q.get('Question_Text_Full', ''))
            ws.cell(row=row_idx, column=5, value=q.get('Question_Text_Condensed', ''))
            ws.cell(row=row_idx, column=6, value=q.get('Data_Collection_Guidance', ''))
            ws.cell(row=row_idx, column=7, value=q.get('Reference_Raw', ''))

            cfr_list = q.get('Reference_CFR_List', [])
            ws.cell(row=row_idx, column=8, value='; '.join(cfr_list) if isinstance(cfr_list, list) else cfr_list)

            faa_list = q.get('Reference_FAA_Guidance_List', [])
            ws.cell(row=row_idx, column=9, value='; '.join(faa_list) if isinstance(faa_list, list) else faa_list)

            other_list = q.get('Reference_Other_List', [])
            ws.cell(row=row_idx, column=10, value='; '.join(other_list) if isinstance(other_list, list) else other_list)

            ws.cell(row=row_idx, column=11, value=q.get('PDF_Page_Number', ''))
            ws.cell(row=row_idx, column=12, value=q.get('PDF_Element_Block_ID', ''))

            notes = q.get('Notes', [])
            ws.cell(row=row_idx, column=13, value='; '.join(notes) if isinstance(notes, list) else notes)

        # Adjust column widths
        column_widths = [12, 12, 15, 80, 40, 50, 30, 30, 30, 20, 15, 20, 40]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
